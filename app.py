import os
import sys
from flask import Flask, request, jsonify
from flask_cors import CORS
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
from rake_nltk import Rake
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import openpyxl
import re
import json
from tqdm import tqdm
import openai

# Ensure UTF-8 encoding
sys.stdout.reconfigure(encoding='utf-8')

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get('SECRET_KEY', 'your_fallback_secret_key')

openai.api_key = os.environ.get('OPENAI_API_KEY', 'sk-proj-PG9DppZOfijgZR9Bq4LtT3BlbkFJfSe5FhdkGrvX4TGDO8Fa')

nltk.download('punkt')
nltk.download('stopwords')

additional_stopwords = [
    # English stopwords
    'students', 'student', 'course', 'courses', 'module', 'modules', 'lecturer', 'lecturers',
    'teacher', 'teachers', 'lecture', 'lectures', 'academic', 'study', 'studies', 'degree',
    'degrees', 'program', 'programs', 'syllabus', 'syllabi', 'curriculum', 'curricula',
    'knowledge', 'understand', 'understanding', 'skill', 'skills', 'competence',
    'competencies', 'learning', 'outcome', 'outcomes', 'objective', 'objectives', 'aim',
    'aims', 'goal', 'goals', 'learn', 'apply', 'develop', 'gain', 'achieve', 'explore',
    'examine', 'analyze', 'evaluate', 'undergo', 'include', 'includes', 'including',
    'provide', 'provides', 'provided', 'offer', 'offering', 'offered', 'cover', 'covers',
    'covered', 'require', 'requires', 'required', 'assess', 'assesses', 'assessed',
    'assessment', 'broad', 'basic', 'relevant', 'important', 'simple', 'advanced',
    'completed', 'acquire', 'reflect', 'reflecting', 'ensure', 'describe', 'explain',
    'discuss', 'create', 'master', 'calculate', 'familiarize', 'concepts', 'methods',
    'theories', 'problems', 'results', 'arguments', 'conclusions', 'answers', 'reader',
    'assumptions', 'models', 'connections', 'subjects', 'topics', 'tools', 'data',
    'information', 'technology', 'systems', 'techniques', 'solutions', 'challenges',
    'applications', 'strategies', 'completing', 'end', 'overview', 'various',
    
    # Norwegian stopwords
    'studentene', 'kunnskap', 'forstå', 'ulike', 'kurset', 'ferdigheter', 'kandidater',
    'kandidat', 'inkludert', 'adressere', 'prosesser', 'kunne', 'øve', 'organisasjon', 'exc',
    'ele', 'bmp', 'gra', 'bik', 'bst', 'smc', 'slm', 'man', 'dre', 'fork', 'ent', 'bøk',
    'lus', 'jur', 'fak', 'met', 'edi', 'eba', 'fin', 'kls', 'str', 'bth', 'mrk', 'ems',
    'bin', 'dig', 'mad', 'nsa', 'modul', 'moduler', 'kurs', 'kursene', 'studie', 'studier',
    'grader', 'program', 'programmer', 'pensum', 'pensumet', 'pensumliste', 'læreplan',
    'læreplaner', 'kompetanse', 'kunnskaper', 'mål', 'målsetting', 'målene', 'utvikle',
    'oppnå', 'utforske', 'undersøke', 'analysere', 'evaluere', 'gjennomgå', 'inkluderer',
    'inkludering', 'gi', 'tilby', 'tilbyr', 'dekke', 'dekker', 'dekking', 'krever', 'krevet',
    'vurdere', 'vurderer', 'vurdert', 'bred', 'grunnleggende', 'relevant', 'viktig', 'enkel',
    'avansert', 'fullført', 'skaffe', 'reflektere', 'beskrive', 'forklare', 'diskutere',
    'skape', 'mestere', 'beregne', 'kjent', 'begreper', 'metoder', 'teorier', 'problemer',
    'resultater', 'argumenter', 'konklusjoner', 'svar', 'leser', 'antagelser', 'modeller',
    'forbindelser', 'emner', 'temaer', 'verktøy', 'data', 'informasjon', 'teknologi',
    'systemer', 'teknikker', 'løsninger', 'utfordringer', 'applikasjoner', 'strategier',
    'fullføring', 'slutt', 'oversikt', 'ulike',
    
    # Existing stopwords from your list
    'students', 'student', 'course', 'studentene', 'able', 'kunnskap', 'knowledge', 'understand',
    'ulike', 'understanding', 'different', 'kurset', 'apply', 'skills', 'candidates', 'candidate',
    'end', 'understands', 'forstår', 'able.', 'address', 'including', 'processes', 'kunne', 'practice',
    'øve', 'org', 'exc', 'ele', 'bmp', 'gra', 'bik', 'bst', 'smc', 'slm', 'man', 'dre', 'fork', 'ent',
    'bøk', 'lus', 'jur', 'fak', 'met', 'edi', 'eba', 'fin', 'kls', 'str', 'bth', 'mrk', 'ems', 'bin',
    'dig', 'mad', 'nsa', 'module', 'modul'
]

EXCEL_FILE_PATH = os.path.join('data', 'Kombifil - med ekstern.xlsx')
JSON_FILE_PATH = os.path.join('data', 'courses.json')
CACHE_FILE_PATH = os.path.join('data', 'cache.json')

# Global variable for caching
cache = {}

def remove_stopwords(text, languages=['english', 'norwegian']):
    stop_words = set()
    for lang in languages:
        stop_words.update(set(stopwords.words(lang)))
        
    stop_words.update(additional_stopwords)

    word_tokens = word_tokenize(text)
    filtered_text = ' '.join(w for w in word_tokens if w.lower() not in stop_words and not re.search(r'\b\d+\b', w))
    return filtered_text

fields_to_include = ['Kurskode', 'Kursnavn', 'Learning outcome - Knowledge', 'Learning outcome - Skills', 'Learning outcome - General Competence', 'Course content']

def remove_stopwords_and_extract_keywords(course_data):
    course_code = course_data['Kurskode']
    if course_code in cache:
        return cache[course_code]['filtered_text'], cache[course_code]['keywords']

    combined_info = ' '.join(str(course_data.get(field, '')) for field in fields_to_include).strip()
    filtered_text = remove_stopwords(combined_info)
    r = Rake()
    r.extract_keywords_from_text(filtered_text)
    keywords = ', '.join(r.get_ranked_phrases())

    # Update cache
    cache[course_code] = {'filtered_text': filtered_text, 'keywords': keywords}
    with open(CACHE_FILE_PATH, 'w') as f:
        json.dump(cache, f)

    return filtered_text, keywords

def load_courses_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    courses = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for row in sheet.iter_rows(min_row=2):
        course_data = {headers[i]: cell.value for i, cell in enumerate(row) if headers[i]}
        filtered_text, keywords = remove_stopwords_and_extract_keywords(course_data)
        course_data['FilteredText'] = filtered_text
        course_data['Keywords'] = keywords
        courses.append(course_data)
    return courses

def store_courses_to_json(courses, json_file):
    with open(json_file, 'w') as f:
        json.dump(courses, f)

def load_courses_from_json(json_file):
    with open(json_file, 'r') as f:
        courses = json.load(f)
    return courses

def update_courses(existing_courses, new_course_data):
    if new_course_data['Kurskode'] == 'N/A':
        return  # Skip updating for new courses entered through the form

    for course in existing_courses:
        if course['Kurskode'] == new_course_data['Kurskode']:  # Assuming 'Kurskode' is a unique identifier
            # Update the existing entry
            course.update(new_course_data)
            break
    else:
        # If new_course_data does not exist, append it to existing_courses
        existing_courses.append(new_course_data)

def generate_overlap_explanation(course_1, course_2):
    prompt = (
        f"The following two course descriptions have a high overlap score. "
        f"As a professor, please provide a short explanation on why these courses might be overlapping, and if the two courses could be considered mutually exclusive or not, so that a student can have them both in their study plan, or not. "
        f"keeping your explanation within 250 words:\n\n"
        f"Course 1 Description: {course_1}\n\n"
        f"Course 2 Description: {course_2}\n\n"
        f"Explanation:"
    )
    
    response = openai.ChatCompletion.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You are a helpful assistant with expertise in higher education."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=300,
        stop=["\n\n", "Explanation:"]
    )
    
    explanation = response['choices'][0]['message']['content'].strip()
    explanation = ensure_complete_sentence(explanation)
    return explanation

def ensure_complete_sentence(text):
    # Ensure the text ends with a complete sentence
    if text and text[-1] not in ['.', '!', '?']:
        last_period = text.rfind('.')
        if last_period == -1:
            # No period found, return the original text
            return text
        return text[:last_period + 1]
    return text

def normalize_literature_entry(entry):
    return set(entry.replace("Book: ", "").replace("'", "").replace('"', '').strip().lower().split('\n'))

def find_literature_matches(user_input, existing_courses):
    user_titles = normalize_literature_entry(user_input)
    print(f"User titles after normalization: {user_titles}")

    matches = []
    for course in existing_courses:
        if 'Pensum' in course and course['Pensum']:
            course_titles = set(
                map(lambda x: x.strip().lower(), 
                    course['Pensum'].replace("Book: ", "").replace("'", "").split('|'))
            )

            common_titles = user_titles.intersection(course_titles)
            print(f"Common titles found: {common_titles}")
            if common_titles:
                matches.append({
                    'Existing Course Code': course.get('Kurskode', 'N/A'),
                    'Existing Course Name': course.get('Kursnavn', 'Unknown'),
                    'Literature Matches': ' | '.join(f"Book: '{title}'" for title in common_titles)
                })
        else:
            print(f"No 'Pensum' data for course: {course.get('Kursnavn', 'Unknown')}")
            
            matches.append({
                'Existing Course Code': course.get('Kurskode', 'N/A'),
                'Existing Course Name': course.get('Kursnavn', 'Unknown'),
                'Literature Matches': "No 'Pensum' data available"
            })

    print(f"Matches found: {matches}")
    return matches

def check_course_overlap(new_course_data, existing_courses, overlap_threshold=0.25):
    new_course_combined_info_fields = [
        str(new_course_data.get('Kurskode', '')),
        str(new_course_data.get('Kursnavn', '')),
        str(new_course_data.get('Learning outcome - Knowledge', '')),
        str(new_course_data.get('Course content', ''))
    ]
    new_course_combined_info = ' '.join(new_course_combined_info_fields).strip()
    new_course_combined_info = remove_stopwords(new_course_combined_info)

    new_course_embedding = model.encode([new_course_combined_info], show_progress_bar=False)
    existing_courses_embeddings = [course['embedding'] for course in existing_courses]
    
    cosine_sim = cosine_similarity(new_course_embedding, existing_courses_embeddings)
    overlapping_courses = []

    for idx, sim_score in enumerate(cosine_sim[0]):
        if new_course_combined_info == existing_courses[idx]['combined_info']:
            overlapping_courses.append({
                'Existing Course Code': existing_courses[idx].get('Kurskode', 'N/A'),
                'Existing Course Name': existing_courses[idx]['Kursnavn'],
                'Overlap Score (%)': 100,
                'Keywords': 'Exact match',
                'Explanation': generate_overlap_explanation(new_course_combined_info, existing_courses[idx]['combined_info'])
            })
            continue

        if sim_score > overlap_threshold:
            sim_score_percentage = round(sim_score * 100, 2)
            r = Rake()
            r.extract_keywords_from_text(existing_courses[idx]['combined_info'])
            phrases_with_scores = r.get_ranked_phrases_with_scores()
            keywords = ', '.join(phrase for score, phrase in phrases_with_scores if score > 1)
            
            explanation = None
            if sim_score_percentage >= 60:
                explanation = generate_overlap_explanation(new_course_combined_info, existing_courses[idx]['combined_info'])

            overlapping_courses.append({
                'Existing Course Code': existing_courses[idx].get('Kurskode', 'N/A'),
                'Existing Course Name': existing_courses[idx]['Kursnavn'],
                'Overlap Score (%)': sim_score_percentage,
                'Keywords': keywords,
                'Explanation': explanation
            })
    return overlapping_courses


# Initialize SentenceTransformer model globally
model = SentenceTransformer('distiluse-base-multilingual-cased-v2')

def add_embeddings_to_courses(courses):
    for course in tqdm(courses, desc="Generating embeddings"):
        combined_info_fields = [
            str(course.get('Kurskode', '')), 
            str(course.get('Kursnavn', '')),
            str(course.get('Learning outcome - Knowledge', '')),
            str(course.get('Learning outcome - Skills', '')),
            str(course.get('Learning outcome - General Competence', '')),
            str(course.get('Course content', ''))
        ]
        combined_info = ' '.join(combined_info_fields).strip()
        course['combined_info'] = remove_stopwords(combined_info)
        course['embedding'] = model.encode([course['combined_info']], show_progress_bar=False).tolist()[0]
    return courses

# Load the existing courses from the JSON file when the application starts
if os.path.exists(JSON_FILE_PATH):
    existing_courses = load_courses_from_json(JSON_FILE_PATH)
else:
    existing_courses = load_courses_from_excel(EXCEL_FILE_PATH)
    existing_courses = add_embeddings_to_courses(existing_courses)
    store_courses_to_json(existing_courses, JSON_FILE_PATH)

# Load the cache from JSON file
if os.path.exists(CACHE_FILE_PATH):
    with open(CACHE_FILE_PATH, 'r') as f:
        cache = json.load(f)

@app.route('/', methods=['GET', 'POST']) 
def home():
    if request.method == 'POST':
        new_course_name = request.form['new_course_name']
        learning_outcomes_and_content = request.form['learning_outcomes_and_content']
        literature = request.form.get('literature', '')  # Retrieve literature from form, if any
        
        print("Received literature data:", literature)

        new_course_details = {
            'Kurskode': 'N/A',  # Assuming Kurskode is not provided in the form
            'Kursnavn': new_course_name,
            'Learning outcome - Knowledge': learning_outcomes_and_content,  # Both learning outcomes and course content
            'Course content': learning_outcomes_and_content  # For consistency, assuming course content same as learning outcomes
        }
        update_courses(existing_courses, new_course_details)
        overlapping_courses = check_course_overlap(new_course_details, existing_courses)

        literature_matches = []
        if literature:
            literature_input = "\n".join([line.strip() for line in literature.splitlines() if line.strip()])
            literature_matches = find_literature_matches(literature_input, existing_courses)

        additional_columns = [
            'Kurskode','Kurskode2','Academic Coordinator', 'School','Credits','Undv.språk',
            'Gj.føring','LINK EN','LINK NB','Level of study','Portfolio','Associate Dean',
            'Ansvarlig institutt','Ansvarlig område',
        ]
        additional_info = [{column: course.get(column, '') for column in additional_columns} for course in existing_courses]

        print("Literature Matches:", literature_matches)
        return jsonify({
            'overlapping_courses': overlapping_courses,
            'literature_matches': literature_matches,
            'additional_info': additional_info
        })

    else:
        return jsonify({'message': 'Ready for overlap checking'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
